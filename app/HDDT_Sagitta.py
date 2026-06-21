import os, sys
import re
import csv
import requests
import base64
from datetime import datetime
from bs4 import BeautifulSoup
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.exceptions import RefreshError
from datetime import datetime, date

# Đặt gần đầu file (chung với import)
ALLOWED_EXT = {'.pdf', '.xml', '.inv', '.rar','zip'}
IMAGE_MIME_PREFIX = 'image/'  # bỏ qua mọi loại ảnh

def get_app_dir():
    """
    Trả về thư mục đặt ứng dụng:
    - Khi chạy .exe (PyInstaller): là thư mục chứa file .exe
    - Khi chạy .py (dev): là thư mục chứa file .py
    """
    if getattr(sys, 'frozen', False):  # đang chạy exe đã đóng gói
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))



# ==== (TÙY CHỌN) Excel mode ====
# Nếu dùng Code="ALL" hoặc chạy theo Code từ Excel, cần pandas + openpyxl
try:
    import pandas as pd
    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

SCOPES = ['https://www.googleapis.com/auth/gmail.modify']
TOKEN_FILE = 'tokenSG.json'
CREDS_FILE = 'credentials_SG.json'

# ========= AUTH (TỰ LÀM MỚI TOKEN) =========
def save_token(creds):
    with open(TOKEN_FILE, 'w', encoding='utf-8') as f:
        f.write(creds.to_json())

def load_token():
    if os.path.exists(TOKEN_FILE):
        return Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    return None

def interactive_login():
    flow = InstalledAppFlow.from_client_secrets_file(CREDS_FILE, SCOPES)
    creds = flow.run_local_server(port=0, access_type='offline', prompt='consent')
    save_token(creds)
    return creds

def get_service():
    creds = load_token()
    if not creds:
        creds = interactive_login()
    else:
        if not creds.valid:
            try:
                if creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                    save_token(creds)
                else:
                    creds = interactive_login()
            except RefreshError:
                creds = interactive_login()
    return build('gmail', 'v1', credentials=creds)

# ========= TIỆN ÍCH =========
def download_file_from_link(url, save_dir, filename):
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        response = requests.get(url, stream=True, headers=headers, timeout=20)
        if response.status_code == 200:
            filepath = os.path.join(save_dir, filename)
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(1024):
                    f.write(chunk)
            print(f"✅ Đã tải: {filename}")
            return filename
        else:
            print(f"⚠️ Không tải được {url} - Mã: {response.status_code}")
    except Exception as e:
        print(f"⚠️ Lỗi khi tải link: {e}")
    return None

def download_attachments(service, message, save_dir):
    """
    Chỉ tải file đính kèm có đuôi .pdf/.xml/.inv/.rar.
    Bỏ qua toàn bộ phần đính kèm ảnh (mimeType bắt đầu với image/).
    Duyệt đệ quy qua các part lồng nhau.
    """
    payload = message.get('payload', {}) or {}
    parts = payload.get('parts', []) or []

    def allowed_by_name_and_mime(filename: str, mime_type: str) -> bool:
        # Bỏ qua nếu là ảnh
        if mime_type and mime_type.lower().startswith(IMAGE_MIME_PREFIX):
            return False
        # Kiểm tra theo đuôi file
        if filename:
            ext = os.path.splitext(filename)[1].lower()
            if ext in ALLOWED_EXT:
                return True
            # không thuộc whitelist -> bỏ
            return False
        # Không có filename: chỉ chấp nhận nếu mime-type là pdf/xml/rar & có attachmentId
        if mime_type:
            mt = mime_type.lower()
            if any(x in mt for x in ['pdf', 'xml', 'rar']):
                return True
        return False

    def walk_parts(p):
        stack = p[:]
        while stack:
            part = stack.pop()
            # nếu còn layer con, tiếp tục duyệt
            if part.get('parts'):
                stack.extend(part['parts'])
                continue
            yield part

    os.makedirs(save_dir, exist_ok=True)

    has_attachment = False
    attachments = []

    for part in walk_parts(parts):
        filename = part.get('filename') or ''
        mime_type = part.get('mimeType', '') or ''
        body = part.get('body', {}) or {}

        # Chỉ xử lý part có attachmentId (tức là file đính kèm thực sự)
        if 'attachmentId' not in body:
            continue

        # Lọc theo whitelist đuôi file & loại MIME (và loại ảnh)
        if not allowed_by_name_and_mime(filename, mime_type):
            # print(f"⏭ Bỏ qua: {filename or mime_type}")
            continue

        att_id = body['attachmentId']
        att = service.users().messages().attachments().get(
            userId='me', messageId=message['id'], id=att_id
        ).execute()

        file_data = base64.urlsafe_b64decode(att['data'].encode('UTF-8'))

        # Nếu thiếu tên file, thử đoán theo mime
        if not filename:
            if 'pdf' in mime_type.lower():
                filename = 'attachment.pdf'
            elif 'xml' in mime_type.lower():
                filename = 'attachment.xml'
            elif 'rar' in mime_type.lower():
                filename = 'attachment.rar'
            else:
                filename = 'attachment.bin'

        # Tránh ghi đè tên trùng
        out_path = os.path.join(save_dir, filename)
        if os.path.exists(out_path):
            name, ext = os.path.splitext(filename)
            i = 2
            while os.path.exists(os.path.join(save_dir, f"{name}_{i}{ext}")):
                i += 1
            out_path = os.path.join(save_dir, f"{name}_{i}{ext}")

        with open(out_path, 'wb') as f:
            f.write(file_data)

        print(f"📎 Đã lưu file đính kèm: {os.path.basename(out_path)}")
        has_attachment = True
        attachments.append(os.path.basename(out_path))

    return has_attachment, attachments

def extract_invoice_number(soup):
    try:
        patterns = [
            r'Hóa đơn mới số[:：\-]?\s*(\d+)',
            r'Số hóa đơn[:：\-]?\s*(\d+)',
        ]
        all_text = soup.get_text(separator='\n', strip=True)
        for pattern in patterns:
            match = re.search(pattern, all_text, re.IGNORECASE)
            if match:
                return match.group(1)

        # fallback span lân cận
        for tag in soup.find_all(string=True):
            if "hóa đơn" in tag.lower():
                parent = tag.parent
                if parent:
                    for span in parent.find_all("span"):
                        txt = span.get_text(strip=True)
                        if txt.isdigit():
                            return txt

        # fallback: lấy số 6-8 chữ số cuối cùng trong mail
        matches = re.findall(r'\b\d{6,8}\b', all_text)
        if matches:
            return matches[-1]
    except Exception as e:
        print(f"Lỗi extract_invoice_number: {e}")
    return "invoice"

def extract_company_name(soup):
    try:
        for span in soup.find_all("span"):
            if "Tên khách hàng" in span.get_text():
                match = re.search(r'Tên khách hàng[:：\-]?\s*(.+)', span.get_text())
                if match:
                    return match.group(1).strip()
        text = soup.get_text()
        match = re.search(r'Tên khách hàng[:：\-]?\s*(.+)', text)
        if match:
            return match.group(1).strip()
    except:
        pass
    return "Không rõ"

def extract_links(soup):
    pdf_url = xml_url = None
    for a in soup.find_all('a', href=True):
        href = a['href'].lower()
        if 'pdfdownload' in href or href.endswith('.pdf'):
            pdf_url = a['href']
        elif 'getinvoice' in href or href.endswith('.xml'):
            xml_url = a['href']
    return pdf_url, xml_url

def write_log(log_file, row):
    file_exists = os.path.isfile(log_file)
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    with open(log_file, 'a', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(['Số hóa đơn', 'Tên công ty', 'Ngày email', 'Tên file PDF', 'Tên file XML'])
        writer.writerow(row)

def mark_email_as_read(service, message_id):
    service.users().messages().modify(
        userId='me',
        id=message_id,
        body={'removeLabelIds': ['UNREAD']}
    ).execute()

# ========= XỬ LÝ MỖI EMAIL =========
def process_email(service, message, save_dir, log_file):
    os.makedirs(save_dir, exist_ok=True)

    msg_date = message['internalDate']
    date_str = datetime.fromtimestamp(int(msg_date)/1000).strftime('%Y-%m-%d')

    any_saved = False

    # Lấy nội dung HTML/TEXT
    payload = message.get('payload', {})
    parts = payload.get('parts', []) or []
    html_data = None
    text_data = None

    # Duyệt sâu các part để lấy text/html
    stack = parts[:]
    while stack:
        part = stack.pop()
        if part.get('parts'):
            stack.extend(part['parts'])
            continue
        mt = part.get('mimeType', '')
        if mt == 'text/html' and not html_data:
            html_data = part.get('body', {}).get('data')
        elif mt == 'text/plain' and not text_data:
            text_data = part.get('body', {}).get('data')

    content = ""
    if html_data:
        content = base64.urlsafe_b64decode(html_data).decode('utf-8', errors='ignore')
    elif text_data:
        content = base64.urlsafe_b64decode(text_data).decode('utf-8', errors='ignore')

    # Từ khóa để CHỈ tải đính kèm, bỏ qua link
    keywords = ["TAOLAO"]
    found_keyword = any(kw.lower() in content.lower() for kw in keywords)

    # 1) Thử tải đính kèm trước
    has_attach, attachments = download_attachments(service, message, save_dir)
    if has_attach and attachments:
        for att in attachments:
            write_log(
                log_file,
                ["(từ file đính kèm)", "Không rõ", date_str,
                 att if att.lower().endswith('.pdf') else "",
                 att if att.lower().endswith('.xml') else ""]
            )
        any_saved = True

    # # 2) Nếu có từ khóa → chỉ tải đính kèm, không tải link
    # if found_keyword:
    #     print(f"Email chứa từ khóa {keywords}. Chỉ tải tệp đính kèm, bỏ qua link!")
    #     if any_saved:
    #         mark_email_as_read(service, message['id'])
    #     return

    # 3) Nếu không có HTML, kết thúc
    if not html_data:
        if any_saved:
            mark_email_as_read(service, message['id'])
        else:
            print("⚠️ Email không chứa nội dung HTML")
        return

    # 4) Trích link từ HTML
    html_decoded = base64.urlsafe_b64decode(html_data).decode('utf-8', errors='ignore')
    soup = BeautifulSoup(html_decoded, 'html.parser')

    invoice_number = extract_invoice_number(soup)
    company_name = extract_company_name(soup)
    pdf_url, xml_url = extract_links(soup)

    pdf_file = xml_file = ""

    # if pdf_url:
    #     saved = download_file_from_link(pdf_url, save_dir, f"{invoice_number}.pdf")
    #     if saved:
    #         pdf_file = saved
    #         any_saved = True
    # else:
    #     print("⚠️ Không tìm thấy link PDF")
    #
    # if xml_url:
    #     saved = download_file_from_link(xml_url, save_dir, f"{invoice_number}.xml")
    #     if saved:
    #         xml_file = saved
    #         any_saved = True
    # else:
    #     print("⚠️ Không tìm thấy link XML")
    #
    # if pdf_file or xml_file:
    #     write_log(log_file, [invoice_number, company_name, date_str, pdf_file, xml_file])

    if any_saved:
        mark_email_as_read(service, message['id'])

# ========= XỬ LÝ THEO TRUY VẤN =========
def process_invoices(service, query, save_dir, log_file):
    results = service.users().messages().list(userId='me', q=query).execute()
    messages = results.get('messages', [])

    if not messages:
        print("⚠️ Không tìm thấy email nào.")
        return

    for msg in messages:
        msg_data = service.users().messages().get(userId='me', id=msg['id'], format='full').execute()
        process_email(service, msg_data, save_dir, log_file)



def build_query_from_senders(senders, start_date, end_date, only_unread=False):
    sender_conditions = " OR ".join([f"from:{s}" for s in senders])
    q = f"({sender_conditions}) after:{start_date} before:{end_date}"
    if only_unread:
        q += " is:unread"
    return q

# ========= EXCEL MODE =========
def yn_to_bool(val, default=False):
    if pd.isna(val) or val is None:
        return default
    s = str(val).strip().lower()
    return s in ("y", "yes", "true", "1")

def norm_date(val, fallback=None):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return fallback
    try:
        dt = pd.to_datetime(val)
        return dt.strftime("%Y/%m/%d")
    except Exception:
        s = str(val).strip()
        try:
            return pd.to_datetime(s).strftime("%Y/%m/%d")
        except Exception:
            return fallback

def run_all_from_excel(excel_path, code="ALL", default_start=None, default_end=None, group_same_save_dir=True):
    if not EXCEL_AVAILABLE:
        raise RuntimeError("Thiếu pandas/openpyxl. Cài: pip install pandas openpyxl")

    if default_start is None:
        default_start = datetime.today().strftime("%Y/%m/%d")
    if default_end is None:
        default_end = datetime.today().strftime("%Y/%m/%d")

    df = pd.read_excel(excel_path)  # sheet đầu
    # Chuẩn hoá tên cột
    cols_lower = [c.lower() for c in df.columns]
    required = ["code", "email", "savedir"]
    for r in required:
        if r not in cols_lower:
            raise ValueError(f"Thiếu cột bắt buộc trong Excel: {r}")

    rename_map = {}
    for want in ["code", "email", "savedir", "startdate", "enddate", "onlyunread"]:
        # tìm cột khớp (không phân biệt hoa thường)
        for c in df.columns:
            if c.lower() == want:
                rename_map[c] = want
                break
    df = df.rename(columns=rename_map)

    # Lọc theo code (nếu không phải ALL)
    if code and code.upper() != "ALL":
        df = df[df["code"].astype(str).str.upper() == code.upper()]

    # bỏ dòng trống email
    df = df[~df["email"].isna() & (df["email"].astype(str).str.strip() != "")]
    if df.empty:
        print("⚠️ Không có dòng nào phù hợp trong Excel.")
        return

    service = get_service()

    if group_same_save_dir:
        # gộp theo (code, savedir, start, end, only_unread) để 1 lần query nhiều email
        df["StartDate_norm"] = df.get("startdate", None).apply(lambda v: norm_date(v, default_start))
        df["EndDate_norm"]   = df.get("enddate", None).apply(lambda v: norm_date(v, default_end))
        df["OnlyUnread_bool"] = df.get("onlyunread", None).apply(yn_to_bool)

        group_cols = ["code", "savedir", "StartDate_norm", "EndDate_norm", "OnlyUnread_bool"]
        for keys, sub in df.groupby(group_cols, dropna=False):
            _, save_dir, sdate, edate, only_unread = keys
            save_dir = str(save_dir).strip()
            sdate = sdate or default_start
            edate = edate or default_end
            only_unread = bool(only_unread)

            senders = sorted(set(sub["email"].astype(str).str.strip().tolist()))
            if not senders:
                continue

            os.makedirs(save_dir, exist_ok=True)
            log_file = os.path.join(save_dir, "log_hoadon.csv")
            query = build_query_from_senders(senders, sdate, edate, only_unread=only_unread)

            print("────────────────────────────────────────────")
            print("📦 SAVE DIR:", save_dir)
            print("📥 SENDERS :", ", ".join(senders))
            print("📆 RANGE   :", sdate, "→", edate, "(only_unread:", only_unread, ")")
            print("🔎 QUERY   :", query)

            process_invoices(service, query=query, save_dir=save_dir, log_file=log_file)
    else:
        # chạy từng dòng riêng lẻ
        for _, row in df.iterrows():
            email = str(row["email"]).strip()
            save_dir = str(row["savedir"]).strip()
            sdate = norm_date(row.get("startdate", None), default_start)
            edate = norm_date(row.get("enddate", None), default_end)
            only_unread = yn_to_bool(row.get("onlyunread", None), False)

            if not email or not save_dir:
                continue

            os.makedirs(save_dir, exist_ok=True)
            log_file = os.path.join(save_dir, "log_hoadon.csv")
            query = build_query_from_senders([email], sdate, edate, only_unread=only_unread)

            print("────────────────────────────────────────────")
            print("📧 EMAIL   :", email)
            print("📦 SAVE DIR:", save_dir)
            print("📆 RANGE   :", sdate, "→", edate, "(only_unread:", only_unread, ")")
            print("🔎 QUERY   :", query)

            process_invoices(service, query=query, save_dir=save_dir, log_file=log_file)

# ================== EXCEL MODE (C1/C2/C3/C4/C5 + từ dòng 8) ==================
from openpyxl import load_workbook
APP_DIR = get_app_dir()
EXCEL_PATH = os.path.join(APP_DIR, "Tai_Gmail.xlsx")
#EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "TAI_GMAIL.xlsx")  # <<< ĐỔI ĐƯỜNG DẪN FILE EXCEL Ở ĐÂY

def parse_excel_date(val, cell_name=""):
    """
    Trả về 'YYYY/MM/DD' từ giá trị ô Excel.
    - Nếu là datetime/date: giữ nguyên theo lịch (không đổi ngày-tháng).
    - Nếu là chuỗi: ưu tiên format dd/mm/yyyy (hoặc d/m/yyyy). Không đoán theo locale.
    - Không có giá trị hợp lệ -> raise ValueError để dễ bắt lỗi.
    """
    if isinstance(val, (datetime, date)):
        return datetime(val.year, val.month, val.day).strftime("%Y/%m/%d")

    if isinstance(val, str):
        s = val.strip().replace("-", "/")
        # dd/mm/yyyy
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
        if m:
            d, mth, y = map(int, m.groups())
            # kiểm tra hợp lệ ngày tháng (vd: 31/02 -> lỗi)
            try:
                dt = datetime(y, mth, d)
                return dt.strftime("%Y/%m/%d")
            except ValueError:
                raise ValueError(f"Ngày không hợp lệ tại ô {cell_name}: {s}")
        # y/m/d đã đúng -> tôn trọng
        m2 = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
        if m2:
            y, mth, d = map(int, m2.groups())
            try:
                dt = datetime(y, mth, d)
                return dt.strftime("%Y/%m/%d")
            except ValueError:
                raise ValueError(f"Ngày không hợp lệ tại ô {cell_name}: {s}")

    raise ValueError(f"Không đọc được ngày tại ô {cell_name}: {val!r}")

def _yn_to_bool(val, default=False):
    if val is None:
        return default
    s = str(val).strip().lower()
    return s in ("y", "yes", "true", "1", "x")

def _build_query(senders, start_date, end_date, only_unread=False):
    if senders:
        sender_conditions = " OR ".join([f"from:{s}" for s in senders])
        q = f"({sender_conditions}) after:{start_date} before:{end_date}"
    else:
        # ALL mode: không lọc người gửi
        q = f"after:{start_date} before:{end_date}"
    if only_unread:
        q += " is:unread"
    return q

def run_from_excel_layout(excel_path: str):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Không tìm thấy file Excel: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active  # sheet đầu

    # ============== MODE: ALL ==============
    mode = (ws["C1"].value or "").strip().upper()  # "ALL" hoặc "TC"
    if mode == "ALL":
        save_dir = (ws["C2"].value or "").strip()

        # --- DÙNG HÀM MỚI, BẮT LỖI NẾU Ô TRỐNG/SAI ---
        start_date = parse_excel_date(ws["C3"].value, "C3")
        end_date   = parse_excel_date(ws["C4"].value, "C4")

        only_unread = str(ws["C5"].value).strip().lower() in ("y", "yes", "true", "1", "x")

        if not save_dir:
            raise ValueError("Ô C2 (SaveDir) đang trống!")

        # Kiểm tra start <= end
        if start_date > end_date:
            raise ValueError(f"Khoảng ngày không hợp lệ: {start_date} > {end_date}")

        os.makedirs(save_dir, exist_ok=True)
        log_file = os.path.join(save_dir, "log_hoadon.csv")

        # ALL = không lọc người gửi
        query = f"after:{start_date} before:{end_date}" + (" is:unread" if only_unread else "")

        print("──────── Chạy MODE = ALL ────────")
        print("📂 SaveDir :", save_dir)
        print("📆 Range   :", start_date, "→", end_date, "(only_unread:", only_unread, ")")
        print("🔎 Query   :", query)

        service = get_service()
        process_invoices(service, query=query, save_dir=save_dir, log_file=log_file)
        return

    # ============== MODE: TC ==============
    if mode != "TC":
        raise ValueError(f"Ô C1 phải là 'ALL' hoặc 'TC', giá trị hiện tại: '{mode}'")

    print("──────── Chạy MODE = TC (tùy chọn từng dòng từ hàng 8) ────────")
    service = get_service()

    groups = {}
    row = 8
    while True:
        email_cell = ws[f"B{row}"].value
        save_dir_cell = ws[f"C{row}"].value
        start_cell = ws[f"D{row}"].value
        end_cell = ws[f"E{row}"].value
        unread_cell = ws[f"F{row}"].value

        if (email_cell is None and save_dir_cell is None
                and start_cell is None and end_cell is None and unread_cell is None):
            break

        email = (str(email_cell).strip() if email_cell else "")
        save_dir = (str(save_dir_cell).strip() if save_dir_cell else "")

        if not email or not save_dir:
            row += 1
            continue

        # --- ép ngày bằng hàm mới; nếu trống/sai sẽ raise (để bạn sửa Excel cho đúng) ---
        try:
            sdate = parse_excel_date(start_cell, f"D{row}")
            edate = parse_excel_date(end_cell, f"E{row}")
        except ValueError as e:
            print(f"⚠️ Bỏ qua dòng {row}: {e}")
            row += 1
            continue

        if sdate > edate:
            print(f"⚠️ Bỏ qua dòng {row}: khoảng ngày không hợp lệ {sdate} > {edate}")
            row += 1
            continue

        only_unread = str(unread_cell).strip().lower() in ("y", "yes", "true", "1", "x")

        key = (save_dir, sdate, edate, only_unread)
        groups.setdefault(key, set()).add(email)
        row += 1

    if not groups:
        print("⚠️ Không có dòng hợp lệ từ hàng 8 trở xuống.")
        return

    for (save_dir, sdate, edate, only_unread), senders in groups.items():
        senders = sorted(list(senders))
        os.makedirs(save_dir, exist_ok=True)
        log_file = os.path.join(save_dir, "log_hoadon.csv")

        sender_conditions = " OR ".join([f"from:{s}" for s in senders])
        query = f"({sender_conditions}) after:{sdate} before:{edate}" + (" is:unread" if only_unread else "")

        print("────────────────────────────────────────────")
        print("📂 SaveDir :", save_dir)
        print("📥 Senders :", ", ".join(senders))
        print("📆 Range   :", sdate, "→", edate, "(only_unread:", only_unread, ")")
        print("🔎 Query   :", query)

        process_invoices(service, query=query, save_dir=save_dir, log_file=log_file)


# ================== ENTRY POINT ==================
def main():
    print("=== Tải hóa đơn Gmail bắt đầu ===")
    # TODO: gọi logic của bạn ở đây, ví dụ:
    # service = get_service()
    # process_invoices(service, query=query, save_dir=save_dir, log_file=log_file)

    run_from_excel_layout(EXCEL_PATH)

    print("\n✅ Hoàn tất tải hóa đơn!")
    # Nếu người dùng double-click .exe (không chạy từ sẵn CMD),
    # giữ cửa sổ lại để họ đọc thông báo:
    try:
        if getattr(sys, "frozen", False) and not sys.stdout.isatty():
            input("Nhấn Enter để thoát...")
    except Exception:
        pass

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Lỗi: {e}")
        try:
            if getattr(sys, "frozen", False) and not sys.stdout.isatty():
                input("Nhấn Enter để thoát...")
        except Exception:
            pass
        raise